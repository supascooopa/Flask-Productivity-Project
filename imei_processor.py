from openpyxl import Workbook, load_workbook
from imei_db_api import imei_finder
from datetime import datetime
from io import BytesIO

def imei_machine(work_book):
    # --- Loading workbook --- #
    wb = load_workbook(work_book)
    sheet_names_lst = wb.sheetnames

    # --- Creating new workbook ---#
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(["IMEI 1", "IMEI 2", "BRAND", "MODEL"])
    row_number = 2

    # going through each page
    for sheet_names in sheet_names_lst:
        ws = wb[sheet_names]
        # --- Iterating over all the columns --- #
        for cols in ws.iter_cols(values_only=True):
            # Assigning the column header that contains the description of the phone
            column_headers = cols[0].strip()
            # Cleaning up the column tuple from None types
            clean_cols = [c_cols for c_cols in cols if c_cols is not None]
            # Usually phone descriptions end with 1 or 2 so here we filter for phones essentially
            if column_headers is not None and column_headers.endswith("1"):
                # Assigning the ONLY IMEI to be put through the IMEI info API
                first_imei = int(cols[1])
                print(first_imei)
                imei_data = imei_finder(first_imei)
                for cells in clean_cols:
                    # Checking to see if the cell only contains numbers
                    if isinstance(cells, float):
                        # Appending to the new ws and leaving the second IMEI slot empty
                        new_ws.append([cells, " ", imei_data[1], imei_data[0].upper()])
            # filtering for the second IMEI
            elif column_headers is not None and column_headers.endswith("2"):
                for cells in cols:
                    if isinstance(cells, float):
                        # Assigning the second IMEI next to the first one we put into the ws above.
                        new_ws.cell(row=row_number, column=2, value=cells)
                        # Need to keep track of number of rows the s
                        row_number += 1
    now = datetime.now().strftime("%d.%m.%Y")
    file_name = "static/files/"+"imei" + now + ".xlsx"
    new_wb.save(file_name)
    virtual_wb = BytesIO()
    new_wb.save(virtual_wb)
    virtual_wb.seek(0)
    return virtual_wb



