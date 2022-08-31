import openpyxl
from flask import Flask, Response
from flask import render_template
from wtforms import FileField, SubmitField, StringField, FieldList, FormField, PasswordField
from flask_wtf import FlaskForm
from werkzeug.utils import secure_filename
import os
from wtforms.validators import InputRequired, length
from flask import send_from_directory, abort
from flask import send_file
from flask import url_for
from flask import redirect
from flask import request
from whatsapp import text_parser_ctwo
# from company_1 import company_number_one
from openpyxl import load_workbook
import io
from datetime import datetime
from flask_sqlalchemy import SQLAlchemy
from flask_login import UserMixin, login_user, LoginManager, login_required, logout_user, current_user
from flask_bcrypt import Bcrypt
from web_form_automator import emo_automator
from imei_processor import imei_machine
from s3_file_manager import file_downloader, file_uploader, file_getter


app = Flask(__name__)
# app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL').replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config["SECRET_KEY"] = os.environ.get("secret_key")
app.config["UPLOAD_FOLDER"] = os.path.abspath("static\\files")  # configure upload folder
app.config["MAX_CONTENT_LENGTH"] = 400 * 1024  # Maximum upload file size is 400kb
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)


# ---- CONFIGURING LOGIN MANAGER --- #
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login_page"


class User(db.Model, UserMixin):
    """ Database where the user information will be stored """
    id = db.Column(db.Integer, primary_key=True)
    user_name = db.Column(db.String(10), nullable=False, unique=True)
    user_password = db.Column(db.String(128), nullable=False)

# db.create_all()


class UserLogin(FlaskForm):
    user_name_field = StringField("User Name", validators=[InputRequired(), length(min=2, max=20)])
    password_field = PasswordField("Password", validators=[InputRequired(), length(min=8, max=20)])
    submit = SubmitField("Login")


class UserRegistration(FlaskForm):
    user_name_field = StringField("User Name", validators=[InputRequired(), length(min=2, max=20)])
    password_field = PasswordField("Password", validators=[InputRequired(), length(min=8, max=20)])
    submit = SubmitField("Submit")


class UploadFileForm(FlaskForm):
    file_field = FileField("File", validators=[InputRequired()])  # file upload field
    submit = SubmitField("Upload")


class PhoneColorEntry(FlaskForm):
    color_entry_field = StringField("color", default="-")  # string field for phone colors with " - " as default
    number_entry_field = StringField("quantity", default="-")  # string field for phone quantity with " - " as default


class POAutomationForm(FlaskForm):
    phones = FieldList(FormField(PhoneColorEntry))  # In order to get a dynamic form, I used FieldList option
    submit = SubmitField("Submit")


def writing_to_memory(path):
    """A function to write files to memory and delete the created file"""

    file_path = os.path.abspath(f"static/files/{path}")  # creating file path
    return_data = io.BytesIO()  # ???
    with open(file_path, "rb") as byte_file:  # reading as byte 'rb'
        return_data.write(byte_file.read())  # writing to memory
    return_data.seek(0)  # returning the cursor to the beginning
    os.remove(file_path)  # removing the saved file
    return return_data


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


@app.route("/", methods=['GET', 'POST'])
def login_page():
    form = UserLogin()
    if form.validate_on_submit():
        user = User.query.filter_by(user_name=form.user_name_field.data).first()
        if user:
            if bcrypt.check_password_hash(user.user_password, form.password_field.data):
                login_user(user)
        return redirect(url_for("home"))
    return render_template("login.html", login_form=form)


@app.route('/logout', methods=['GET', 'POST'])
@login_required
def logout():
    logout_user()
    return redirect(url_for('login_page'))


@app.route("/register-user", methods=['GET', 'POST'])
@login_required
def registration_page():
    cu_id = current_user.id
    if cu_id == 1:
        form = UserRegistration()
        if form.validate_on_submit():
            hashed_password = bcrypt.generate_password_hash(form.password_field.data)
            new_user = User(user_name=form.user_name_field.data, user_password=hashed_password)
            db.session.add(new_user)
            db.session.commit()
            return redirect(url_for("login_page"))
        return render_template("register.html", register_form=form)


@app.route("/home")
@login_required
def home():
    """ Home Page
     Where links to various parts of the application are located"""
    return render_template("index.html")


@app.route("/texttocsv", methods=["GET", "POST"])
@login_required
def text_to_csv():
    form = UploadFileForm()
    if form.validate_on_submit():
        file = form.file_field.data  # grab the file
        binary_data = file.read()  # converts the file to binary
        secure_file_name = secure_filename(file.filename)
        file_name = secure_file_name.split(".")[0]
        # Below function parser a text file (Specifically Whatsapp price list sent by the company two)
        # in order to convert it to a csv file we pass the directory location of the file we want to be converted
        modified_file = text_parser_ctwo(binary_data.decode("utf-8"))
        # In order send the user the modified data we split the name function gave back to us
        # and take the modified file name and redirect that name to another function
        return Response(modified_file, mimetype="application/csv",
                        headers={"Content-Disposition": f"attachment;filename={file_name}.csv"}
                        )
    return render_template("upload.html", form=form)


@app.route("/pdftoexcel", methods=["GET", "POST"])
@login_required
def pdf_to_excel():
    form = UploadFileForm()
    message = "Only add the last three pages of the price list! It can't be a image based PDF!!"
    if form.validate_on_submit():
         # same stuff like the above
        # TODO put the grabbing and saving file in a function
    #     file = form.file_field.data  # grab the file
    #     secure_file_name = secure_filename(file.filename)
    #     file.save(os.path.join(os.path.abspath(os.path.dirname(__file__)),
    #                            app.config["UPLOAD_FOLDER"],
    #                            secure_file_name))  # save the file
    #     modified_file_name = company_number_one(os.path.abspath(f"static\\files\\{secure_file_name}"))
    #     new_file_name = modified_file_name.split("\\")[-1]
    #     return redirect(url_for("excel_download", file_name=new_file_name))
        return "hello stranger!"
    return render_template("upload.html", form=form, message=message)


@app.route("/imei-machine", methods=["GET", "POST"])
@login_required
def imei_automator():
    form = UploadFileForm()
    if form.validate_on_submit():
        # --- For local testing --- #
        # file = form.file_field.data  # grab the file
        # secure_file_name = secure_filename(file.filename)
        # file.save(os.path.join(os.path.abspath(os.path.dirname(__file__)),
        #                        app.config["UPLOAD_FOLDER"],
        #                        secure_file_name))
        # modified_excel_file = imei_machine(os.path.abspath(f"static\\files\\{secure_file_name}"))
        # file_name = modified_excel_file.split("/")[-1]
        file = form.file_field.data
        binary_data = file.read()  # converts the file to binary
        secure_file_name = secure_filename(file.filename)
        # bucket_key = f"excel_files/{secure_file_name}"
        # file_uploader(binary_data, bucket_key)
        # excel_file = io.BytesIO(file_getter(bucket_key))
        excel_file = io.BytesIO(binary_data)
        modified_file = imei_machine(excel_file)
        return Response(modified_file,
                        mimetype="application/vnd.ms-excel",
                        headers={"Content-Disposition": f"attachment;filename={secure_file_name}"})
    return render_template("upload.html", form=form)


@app.route("/excel-download/<file_name>")
@login_required
def excel_download(file_name):
    # same as the above function except it works with pdf_to_excel function
    try:
        return_data = writing_to_memory(file_name)
        return send_file(return_data, mimetype="application/vnd.ms-excel", download_name=file_name)
    except FileNotFoundError:
        abort(404)


@app.route("/automation-sheet-download")
@login_required
def automation_sheet():
    try:
        # FOR TESTING
        # return send_from_directory(os.path.abspath("static\\browser_automation\\XL"), file_name, as_attachment=True)
        # FOR DEPLOYMENT
        # --- Fetching the file from s3 --- #
        file = file_downloader("browser_automation/excel_sample/excel_sample.xlsx")
        # --- returning file as a response --- #
        return Response(file["Body"].read(),
                        mimetype="application/vnd.ms-excel",
                        headers={"Content-Disposition": f"attachment;filename=excel_sample.xlsx"})
    except FileNotFoundError:
        abort(404)


@app.route("/PO-excel-upload", methods=["GET", "POST"])
@login_required
def upload_excel_file():
    form = UploadFileForm()
    message = "Please follow the example table below when trying to upload an excel sheet!"
    route_info = request.url_rule
    if form.validate_on_submit():
        file = form.file_field.data  # grab the file
        secure_file_name = secure_filename(file.filename)
        binary_data = file.read()
        bucket_key = f"excel_files/{secure_file_name}"
        file_uploader(binary_data, bucket_key)
        # file.save(os.path.join(os.path.abspath(os.path.dirname(__file__)), app.config["UPLOAD_FOLDER"],
                               # secure_file_name))
        return redirect(url_for("po_automator", excel_file_name=secure_file_name))
    return render_template("upload.html", form=form, message=message, route=route_info.rule)


@app.route("/PO-automator/<excel_file_name>", methods=["GET", "POST"])
@login_required
def po_automator(excel_file_name):
    # opens the workbook from the specified directory
    bucket_key = f"excel_files/{excel_file_name}"
    data = file_getter(bucket_key)
    binary_data = io.BytesIO(data)
    wb = load_workbook(binary_data)
    # activating worksheet
    ws = wb.active
    # grabbing values by row
    rows = [list(rows) for rows in ws.iter_rows(values_only=True)]
    # creating a dictionary and inserting into FieldList
    form_dict = [{i[0]:i[1]} for i in rows]
    po_form = POAutomationForm(phones=form_dict)
    if po_form.validate_on_submit():
        # using zip() to go over the FieldList fields and active sheet rows
        for data, phones in zip(po_form.data["phones"], rows):
            # grabbing color and quantity data from fields data
            color = data["color_entry_field"]
            quantity = data['number_entry_field']
            # appending data into rows
            phones.extend([color, quantity])
        new_list = []
        for phones in rows:
            # Checking to see if user entered color or quantity
            if "-" not in phones:
                # splitting colors and quantities
                phone_color = phones[-2].split(',')
                phone_quantity = phones[-1].split(',')
                phone_model = phones[0]
                phone_description = phones[1]
                phone_price = phones[2]
                # had to apply a counter here in order to access quantities
                counter = 0
                for color in phone_color:
                    new_name = [phone_model+" "+phone_description+" "+color, phone_quantity[counter], phone_price]
                    new_list.append(new_name)
                    counter += 1
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        # adding colum headers
        new_ws.append(["Phone description", "Phone quantity", "Price"])
        # appending the data into rows
        for rows in new_list:
            new_ws.append(rows)
        # choosing name for file and saving the excel file
        virtual_wb = io.BytesIO()
        new_wb.save(virtual_wb)
        virtual_wb.seek(0)
        return Response(virtual_wb,
                        mimetype="application/vnd.ms-excel",
                        headers={"Content-Disposition": f"attachment;filename={excel_file_name}"})
    return render_template("po_automator_page.html", lst=rows, form=po_form, len=len(rows))


@app.route("/web_form_automator", methods=["GET", "POST"])
@login_required
def form_automator():
    upload_form = UploadFileForm()
    route_info = request.url_rule
    file_name = secure_filename("TEST.xlsx")
    if upload_form.validate_on_submit():
        emo_automator()
        return redirect(url_for("form_automator", message="Your task has been completed"))
    return render_template("upload.html", form=upload_form, route=route_info.rule, file_name=file_name)


@app.errorhandler(413)
def request_entity_too_large(error):
    return "File size larger than 400Kb", 413


if __name__ == "__main__":
    app.run(debug=True)

