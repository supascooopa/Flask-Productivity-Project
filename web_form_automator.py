import os
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import ElementNotInteractableException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service


def auto_fill(dictionary):
    for key in dictionary:
        auto_input = driver.find_element(By.ID, key)
        try:
            if key == "ctl00_cp_txt_txt_firma_e_posta_text":
                auto_input.clear()
                auto_input.send_keys(dictionary[key])
                auto_input.send_keys(Keys.ENTER)
                time.sleep(1)
            else:
                driver.execute_script("arguments[0].click();", auto_input)
                auto_input.clear()
                auto_input.send_keys(dictionary[key])
                auto_input.send_keys(Keys.ENTER)
                time.sleep(1)

        except ElementNotInteractableException:
            auto_input.click()
            time.sleep(3)
            auto_input.send_keys(dictionary[key])


def emo_automator():
    global driver
    service = Service("static\\browser_automation\\geckodriver.exe")
    driver = webdriver.Firefox(service=service)

    excel_file = os.listdir("static\\browser_automation\\XL")[0]
    wb = openpyxl.load_workbook(f"static\\browser_automation\\XL\\{excel_file}")
    ws = wb.active

    list_of_ce = [file for file in os.listdir("static\\browser_automation\\CE")]
    path_to_ce = [os.path.abspath("static\\browser_automation\\CE") + "\\" + file + " " for file in list_of_ce]
    combine = "\n".join(path_to_ce)

    url = "https://www.ktemo.org/e_EMO"

    emo_company_auto_fill = {
        "ctl00_cp_cmb_cmb_ithalatci_ad_liste_combo_Input": "blue",
        "ctl00_cp_txt_txt_vergi_no_text": "",
        "ctl00_cp_msktxt_msktxt_firma_tel_text": "",
        "ctl00_cp_txt_txt_firma_e_posta_text": "",
        "ctl00_cp_txt_txt_ihracatci_ad_text": "",
        # You should inspect elements after you click them because the thing you are looking for might not be visible!
        "ctl00_cp_cmb_cmb_gonderilen_adres_combo_Input": "",
        "ctl00_cp_txt_txt_p_f_no_text": "",
        "ctl00_cp_txt_txt_fatura_tarihi_text": "",
        "ctl00_cp_cmb_cmb_para_birimi_combo_Input": "USD"

    }

    exporter_name = ws["B3"].value
    emo_company_auto_fill["ctl00_cp_txt_txt_ihracatci_ad_text"] = exporter_name
    export_address = ws["C3"].value
    emo_company_auto_fill["ctl00_cp_cmb_cmb_gonderilen_adres_combo_Input"] = export_address
    invoice_number = ws["D3"].value
    emo_company_auto_fill["ctl00_cp_txt_txt_p_f_no_text"] = invoice_number
    currency_type = ws["E3"].value
    emo_company_auto_fill["ctl00_cp_cmb_cmb_para_birimi_combo_Input"] = currency_type
    invoice_date = ws["F3"].value
    emo_company_auto_fill["ctl00_cp_txt_txt_fatura_tarihi_text"] = invoice_date
    tax_number = ws["G3"].value
    emo_company_auto_fill["ctl00_cp_txt_txt_vergi_no_text"] = tax_number
    phone_number = ws["H3"].value
    emo_company_auto_fill["ctl00_cp_msktxt_msktxt_firma_tel_text"] = phone_number
    email_address = ws["I3"].value
    emo_company_auto_fill["ctl00_cp_txt_txt_firma_e_posta_text"] = email_address

    emo_products_auto_fill = {
        "ctl00_cp_cmbCol_grd_000_cmb_malzeme_tipi_combo_Input": "2",
        "ctl00_cp_txtCol_grd_malzeme_listesi_txtCol_urun_aciklamasi_text": "",
        "ctl00_cp_txtCol_grd_malzeme_listesi_txtCol_marka_text": "",
        "ctl00_cp_txtCol_grd_malzeme_listesi_txtCol_Miktar_text": "",
        "ctl00_cp_cmbCol_grd_malzeme_listesi_cmbCol_birim_combo_Input": "Adet",
        "ctl00_cp_txtCol_grd_malzeme_listesi_txtCol_Birim_Fiyat_text": "101"
    }

    # ----------------ENTERING SYSTEM & FILLING OUT THE FORM----------------- #

    driver.get(url)
    login = driver.find_element(By.CSS_SELECTOR, "td a")
    login.click()
    time.sleep(5)
    auto_fill(emo_company_auto_fill)

    # ----------------UPLOADING INVOICE----------------- #

    upload_invoice = driver.find_element(By.ID, "ctl00_cp_file_file_proforma_fatura_imgPopupRhea")
    upload_invoice.click()
    driver.implicitly_wait(5)
    iframe = driver.find_element(By.XPATH, '//*[@id="modal-body-iframe-0"]')
    driver.switch_to.frame(iframe)
    upload_invoice_button = driver.find_element(By.CLASS_NAME, "ruFileInput")
    pdf_file = os.listdir("static\\browser_automation\\invoice")[0]
    upload_invoice_button.send_keys(os.path.abspath("static\\browser_automation\\invoice") + f"\\{pdf_file}")
    time.sleep(3)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "ctl00_cp_pnlBar1_i0_btnSelect_Upload"))).click()
    driver.switch_to.default_content()

    # ----------------INPUTTING PRODUCTS----------------- #

    time.sleep(3)
    driver.find_element(By.ID, "ctl00_cp_grd_grd_malzeme_listesi_ctl00_ctl02_ctl00_InitInsertButton").click()
    time.sleep(3)
    iframe_products = driver.find_element(By.XPATH, '//*[@id="modal-body-iframe-1"]')
    driver.switch_to.frame(iframe_products)
    for row in ws.iter_rows(min_col=2, max_col=6, min_row=6, values_only=True):
        if None in row:
            pass
        else:
            product_name = row[1]
            emo_products_auto_fill["ctl00_cp_txtCol_grd_malzeme_listesi_txtCol_urun_aciklamasi_text"] = product_name
            product_brand = row[2]
            emo_products_auto_fill["ctl00_cp_txtCol_grd_malzeme_listesi_txtCol_marka_text"] = product_brand
            product_amount = row[3]
            emo_products_auto_fill["ctl00_cp_txtCol_grd_malzeme_listesi_txtCol_Miktar_text"] = product_amount
            product_price = str(row[4])
            emo_products_auto_fill["ctl00_cp_txtCol_grd_malzeme_listesi_txtCol_Birim_Fiyat_text"] = product_price
            auto_fill(emo_products_auto_fill)
            time.sleep(2)
            driver.find_element(By.ID, "ctl00_cp_btnUpdate").click()
            time.sleep(5)
    driver.find_element(By.ID, "ctl00_cp_btnCancel").click()
    driver.switch_to.default_content()

    # ----------------UPLOADING CERTIFICATES----------------- #

    time.sleep(3)
    driver.find_element(By.ID, "ctl00_cp_file_file_002_imgPopupRhea").click()
    time.sleep(3)
    iframe_certificates = driver.find_element(By.XPATH, '//*[@id="modal-body-iframe-2"]')
    driver.switch_to.frame(iframe_certificates)
    upload_certificate_button = driver.find_element(By.CLASS_NAME, "ruFileInput")
    upload_certificate_button.send_keys(combine)
    time.sleep(3)
    unordered_list = driver.find_element(By.ID, "ctl00_cp_pnlBar1_i0_radFileUploadListContainer")
    list_of_li = unordered_list.find_elements(By.TAG_NAME, "li")
    last_il = list_of_li[-1]
    wait = WebDriverWait(driver, 600)
    # below waits until the last list item is uploaded
    wait.until(lambda d: "ruFileLI ruUploading" != last_il.get_attribute('class'))
    driver.find_element(By.ID, "ctl00_cp_pnlBar1_i0_btnSelect_Upload").click()
    driver.switch_to.default_content()


